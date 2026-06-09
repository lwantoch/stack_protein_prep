"""Cell styling and formatting helpers for the XLSX exporter.

All symbols are re-exported from pipeline_xlsx via ``from ._xlsx_formatters import *``.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from stack_protein_preparation._xlsx_columns import (
    # Style constants
    BODY_FILL,
    CENTER_ALIGNMENT,
    DATA_ALIGNMENT,
    DATA_FONT,
    FAIL_FILL,
    FINAL_MODEL_YES_FONT,
    FONT_NAME,
    FONT_SIZE,
    GOLD_HEX,
    GRID_BORDER,
    HEADER_ALIGNMENT,
    HEADER_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    LINK_FONT,
    METALS_MODEL_READY_COLUMN_NAME,
    MUTED_FONT,
    PDB_MISSING_FONT,
    REVIEW_FILL,
    ROW_STRIPE_FILL,
    SKIP_FILL,
    STRONG_REVIEW_FILL,
    SUCCESS_FILL,
    WHITE_BOLD_FONT,
    # Column names needed by formatting functions
    FINAL_MODEL_CREATED_COLUMN_NAME,
    HAS_GAPS_COLUMN_NAME,
    HAS_LIGANDS_COLUMN_NAME,
    HAS_METALS_COLUMN_NAME,
    HAS_NONSTANDARD_RESIDUES_COLUMN_NAME,
    METALS_GEOMETRY_FOUND_COLUMN_NAME,
    METALS_GEOMETRY_MATCH_COLUMN_NAME,
    METALS_GEOMETRY_PROBABLE_COLUMN_NAME,
    PARAMETER_AUDIT_REQUIRES_METAL_COLUMN_NAME,
    PARAMETER_AUDIT_REQUIRES_QM_COLUMN_NAME,
    PARAMETER_AUDIT_REQUIRES_REPAIR_COLUMN_NAME,
    PDB_ID_COLUMN_NAME,
    RETAIN_ALPHAFOLD_VARIANT_COLUMN_NAME,
    RETAIN_GAPS_VARIANT_COLUMN_NAME,
    RETAIN_MODELLER_VARIANT_COLUMN_NAME,
)
from stack_protein_preparation.pipeline_state import (
    STATUS_FAILED,
    STATUS_REQUIRED,
    STATUS_SKIPPED,
    STATUS_SUCCESS,
    STATUS_WARNING,
)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _set_white_bold(cell) -> None:
    cell.font = WHITE_BOLD_FONT
    cell.alignment = CENTER_ALIGNMENT


# ---------------------------------------------------------------------------
# Public cell-level formatters
# ---------------------------------------------------------------------------


def apply_status_cell_color(cell, value: str) -> None:
    """Apply status coloring to one cell."""

    normalized_value = str(value).strip().lower()

    if normalized_value == STATUS_SUCCESS:
        cell.fill = SUCCESS_FILL
    elif normalized_value == STATUS_WARNING:
        cell.fill = STRONG_REVIEW_FILL
        _set_white_bold(cell)
    elif normalized_value == STATUS_REQUIRED:
        cell.fill = REVIEW_FILL
    elif normalized_value == STATUS_SKIPPED:
        cell.fill = SKIP_FILL
        cell.font = MUTED_FONT
    elif normalized_value == STATUS_FAILED:
        cell.fill = FAIL_FILL


def apply_semantic_yes_no_color(
    cell,
    column_name: str,
    value: str,
) -> None:
    """Apply value-based colors with one meaning per color."""

    normalized_value = str(value).strip().lower()

    if column_name == FINAL_MODEL_CREATED_COLUMN_NAME:
        cell.alignment = CENTER_ALIGNMENT
        if normalized_value == "yes":
            cell.font = FINAL_MODEL_YES_FONT
        elif normalized_value == "no":
            cell.font = PDB_MISSING_FONT
        return

    if column_name == HAS_LIGANDS_COLUMN_NAME:
        return

    if column_name == HAS_METALS_COLUMN_NAME:
        if normalized_value == "yes":
            cell.fill = REVIEW_FILL
        return

    if column_name == HAS_NONSTANDARD_RESIDUES_COLUMN_NAME:
        if normalized_value == "yes":
            cell.fill = REVIEW_FILL
        return

    if column_name == HAS_GAPS_COLUMN_NAME:
        if normalized_value == "yes":
            cell.fill = REVIEW_FILL
        return

    if column_name in {
        RETAIN_GAPS_VARIANT_COLUMN_NAME,
        RETAIN_MODELLER_VARIANT_COLUMN_NAME,
        RETAIN_ALPHAFOLD_VARIANT_COLUMN_NAME,
    }:
        if normalized_value == "yes":
            cell.fill = SUCCESS_FILL
        return

    if column_name == METALS_MODEL_READY_COLUMN_NAME:
        cell.alignment = CENTER_ALIGNMENT
        if normalized_value == "yes":
            cell.fill = SUCCESS_FILL
        elif normalized_value == "no":
            cell.fill = FAIL_FILL
        elif normalized_value in {"review", "warning", "required"}:
            cell.fill = REVIEW_FILL
        return

    if column_name == METALS_GEOMETRY_MATCH_COLUMN_NAME:
        cell.alignment = CENTER_ALIGNMENT
        if normalized_value == "yes":
            cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=GOLD_HEX)
        elif normalized_value == "no":
            cell.font = PDB_MISSING_FONT
        return

    if column_name in {
        PARAMETER_AUDIT_REQUIRES_REPAIR_COLUMN_NAME,
        PARAMETER_AUDIT_REQUIRES_QM_COLUMN_NAME,
        PARAMETER_AUDIT_REQUIRES_METAL_COLUMN_NAME,
    }:
        if normalized_value == "true":
            cell.fill = FAIL_FILL
        return


def apply_metal_geometry_pair_color(
    worksheet,
    protein_record_list: list[dict[str, str]],
    column_order: list[str],
    *,
    first_data_row_index: int,
) -> None:
    """Color found/probable metal geometry text from the match decision."""

    try:
        found_column_index = column_order.index(METALS_GEOMETRY_FOUND_COLUMN_NAME) + 1
        probable_column_index = column_order.index(METALS_GEOMETRY_PROBABLE_COLUMN_NAME) + 1
    except ValueError:
        return

    for offset, protein_record in enumerate(protein_record_list):
        geometry_match = str(
            protein_record.get(METALS_GEOMETRY_MATCH_COLUMN_NAME, "")
        ).strip().lower()

        if geometry_match == "yes":
            geometry_font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=GOLD_HEX)
        elif geometry_match == "no":
            geometry_font = PDB_MISSING_FONT
        else:
            continue

        row_index = first_data_row_index + offset
        for column_index in (found_column_index, probable_column_index):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.value not in (None, ""):
                cell.font = geometry_font


def style_header_row(
    worksheet,
    column_order: list[str],
    header_row_index: int,
) -> None:
    """Apply professional header styling."""

    for column_index, _column_name in enumerate(column_order, start=1):
        cell = worksheet.cell(row=header_row_index, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER


def apply_base_data_style(
    worksheet,
    *,
    first_data_row_index: int,
) -> None:
    """Apply the consistent body base before semantic colors."""

    for row_index in range(first_data_row_index, worksheet.max_row + 1):
        base_fill = (
            ROW_STRIPE_FILL
            if (row_index - first_data_row_index) % 2 == 1
            else BODY_FILL
        )

        for cell in worksheet[row_index]:
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = GRID_BORDER
            cell.fill = base_fill


def apply_group_body_shading(
    worksheet,
    column_order: list[str],
    *,
    first_data_row_index: int,
) -> None:
    """No-op: body shading is value-based via apply_base_data_style."""

    return


def apply_row_striping(
    worksheet,
    *,
    first_data_row_index: int,
) -> None:
    """No-op: striping is applied in apply_base_data_style."""

    return


def autosize_worksheet_columns(worksheet) -> None:
    """Set readable bounded widths for all worksheet columns."""

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0

        for row_index in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))

        header_value = str(worksheet.cell(row=1, column=column_index).value or "")

        if header_value.endswith("_path") or "directory" in header_value:
            adjusted_width = 12
        elif column_index == 1:
            adjusted_width = 11
        else:
            adjusted_width = min(max(max_length + 2, 10), 28)

        worksheet.column_dimensions[column_letter].width = adjusted_width


def set_worksheet_view_options(
    worksheet,
    *,
    freeze_cell: str,
    auto_filter_range: str,
) -> None:
    """Apply stable worksheet usability defaults."""

    worksheet.freeze_panes = freeze_cell
    worksheet.auto_filter.ref = auto_filter_range
    worksheet.sheet_view.showGridLines = False


def _set_click_hyperlink(cell, target_path: str) -> None:
    """Store the path as hyperlink target, but show only ``open``."""

    if not target_path:
        return

    try:
        target = Path(target_path).expanduser().resolve()
        cell.value = "open"
        cell.hyperlink = target.as_uri()
        cell.font = LINK_FONT
        cell.alignment = CENTER_ALIGNMENT
    except Exception:
        cell.value = target_path


def _set_row_height(worksheet, row_index: int, height: float) -> None:
    worksheet.row_dimensions[row_index].height = height


def mark_missing_model_pdb_ids(
    worksheet,
    protein_record_list: list[dict[str, str]],
    column_order: list[str],
    *,
    first_data_row_index: int,
) -> None:
    """No-op placeholder — visual marking for missing model PDB IDs not yet implemented."""
    pass
