"""Professional XLSX export for FRUTON pipeline state.

The exporter writes the flat FRUTON pipeline records into a compact multi-sheet
workbook. The first worksheet is a decision dashboard: it starts with a compact
logo header, a small KPI strip, and then shows only the columns needed to decide
whether a final model exists, which chemistry flags are relevant, which model
family is available, and which stage failed. Detailed paths and full-state data remain available in later
worksheets for debugging, but they do not dominate the first view.

The visual style uses an IDIS-aligned FRUTON palette rather than default office colors.
Navy (#2F3761) is used for structure, headers, and final-model text accents. Gold (#8A730F) marks review/warning states. A pre-blended soft navy
(#D5D7DF) approximates #2F3761 at 80% transparency over white and is used as the
main table body shade. Gold is used only for the accent row and review signals.
Red (#C00000) is reserved for hard failures and PDB IDs where no final model was
created.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import inch_to_EMU

from stack_protein_preparation.pipeline_state import (
    create_empty_protein_record,
)
from stack_protein_preparation._xlsx_columns import (  # noqa: F401 (re-exported)
    ALIGNMENT_DIRECTORY_COLUMN_NAME,
    AVAILABLE_MODELS_COLUMN_NAME,
    BODY_FILL,
    CENTER_ALIGNMENT,
    CHEMISTRY_COLUMN_NAME_LIST,
    CHEMISTRY_GROUP_COLUMN_NAME_SET,
    CHEMISTRY_GROUP_FILL,
    COMPONENTS_COLUMN_NAME_LIST,
    COMPONENTS_DIRECTORY_COLUMN_NAME,
    DATA_ALIGNMENT,
    DATA_FONT,
    DERIVED_COLUMN_NAME_LIST,
    FAIL_FILL,
    FASTA_DIRECTORY_COLUMN_NAME,
    FASTA_FILES_DONE_COLUMN_NAME,
    FILLER_DIRECTORY_COLUMN_NAME,
    FILLER_MODEL_PATH_COLUMN_NAME,
    FILLER_MODEL_SOURCE_COLUMN_NAME,
    FILLER_STATUS_COLUMN_NAME,
    FINAL_MODEL_CREATED_COLUMN_NAME,
    FINAL_MODEL_PATH_COLUMN_NAME,
    FINAL_MODEL_TYPE_COLUMN_NAME,
    FINAL_MODEL_YES_FONT,
    FONT_NAME,
    FONT_SIZE,
    GAP_GROUP_COLUMN_NAME_SET,
    GAP_GROUP_FILL,
    GAP_SIZES_COLUMN_NAME,
    GAPS_AND_VARIANTS_COLUMN_NAME_LIST,
    GOLD_HEX,
    GRID_BORDER,
    GRID_HEX,
    HEADER_ALIGNMENT,
    HEADER_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    HAS_GAPS_COLUMN_NAME,
    HAS_LIGANDS_COLUMN_NAME,
    HAS_METALS_COLUMN_NAME,
    HAS_NONSTANDARD_RESIDUES_COLUMN_NAME,
    HYPERLINK_COLUMN_NAME_SET,
    INSERTION_CODES_DONE_COLUMN_NAME,
    INTERNAL_CAPPING_INPUT_PATH_COLUMN_NAME,
    INTERNAL_CAPPING_OUTPUT_PATH_COLUMN_NAME,
    INTERNAL_CAPPING_STATUS_COLUMN_NAME,
    KPI_LABEL_FILL,
    KPI_LABEL_FONT,
    KPI_VALUE_FILL,
    KPI_VALUE_FONT,
    LIGHT_GRID_HEX,
    LINK_FONT,
    LOGO_FILENAME_CANDIDATE_LIST,
    LOGO_HEIGHT_IN,
    LOGO_OFFSET_X_IN,
    LOGO_OFFSET_Y_IN,
    LOGO_WIDTH_IN,
    METALL_PARAMS_DIRECTORY_COLUMN_NAME,
    METALL_PARAMS_MANIFEST_PATH_COLUMN_NAME,
    METALL_PARAMS_SITE_COUNT_COLUMN_NAME,
    METALL_PARAMS_STATUS_COLUMN_NAME,
    METALS_CHECK_LOG_PATH_COLUMN_NAME,
    METALS_CHECK_MANIFEST_PATH_COLUMN_NAME,
    METALS_CHECK_STATUS_COLUMN_NAME,
    METALS_CLASS_COLUMN_NAME,
    METALS_GEOMETRY_FOUND_COLUMN_NAME,
    METALS_GEOMETRY_MATCH_COLUMN_NAME,
    METALS_GEOMETRY_PROBABLE_COLUMN_NAME,
    METALS_ION_TYPE_COLUMN_NAME,
    METALS_MODEL_READY_COLUMN_NAME,
    METALS_PARAMETER_REFERENCE_COLUMN_NAME,
    MODEL_GROUP_COLUMN_NAME_SET,
    MODEL_GROUP_FILL,
    MUTED_FONT,
    NAVY_HEX,
    NONSTD_RESIDUE_PARAMS_MANIFEST_PATH_COLUMN_NAME,
    NONSTD_RESIDUE_PARAMS_N_RESIDUES_COLUMN_NAME,
    NONSTD_RESIDUE_PARAMS_STATUS_COLUMN_NAME,
    N_GAPS_COLUMN_NAME,
    OPTIONAL_KNOWN_EXTRA_COLUMN_NAME_LIST,
    ORANGE_FILL,
    ORANGE_HEX,
    OVERVIEW_COLUMN_NAME_LIST,
    PARAMETER_AUDIT_JSON_PATH_COLUMN_NAME,
    PARAMETER_AUDIT_LOG_PATH_COLUMN_NAME,
    PARAMETER_AUDIT_REQUIRES_METAL_COLUMN_NAME,
    PARAMETER_AUDIT_REQUIRES_QM_COLUMN_NAME,
    PARAMETER_AUDIT_REQUIRES_REPAIR_COLUMN_NAME,
    PARAMETER_AUDIT_STATUS_COLUMN_NAME,
    PATHS_COLUMN_NAME_LIST,
    PDB_DIRECTORY_COLUMN_NAME,
    PDB_FONT,
    PDB_ID_COLUMN_NAME,
    PDB_MISSING_FONT,
    PDB_SYNC_DONE_COLUMN_NAME,
    PREPARED_ALPHAFOLD_OUTPUT_PATH_COLUMN_NAME,
    PREPARED_DIRECTORY_COLUMN_NAME,
    PREPARED_GAPS_OUTPUT_PATH_COLUMN_NAME,
    PREPARED_MODELLER_OUTPUT_PATH_COLUMN_NAME,
    PREPARED_OUTPUTS_COLUMN_NAME_LIST,
    PREPARED_STRUCTURE_ACTUAL_RANGE_COLUMN_NAME,
    PREPARED_STRUCTURE_OUTPUT_PATH_COLUMN_NAME,
    PREPARED_STRUCTURE_PROTEIN_INPUT_PATH_COLUMN_NAME,
    PREPARED_STRUCTURE_STATUS_COLUMN_NAME,
    PREPARED_STRUCTURE_VARIANT_COLUMN_NAME,
    PREPARATION_COLUMN_NAME_LIST,
    PROTONATION_INPUT_PATH_COLUMN_NAME,
    PROTONATION_INPUT_SOURCE_COLUMN_NAME,
    PROTONATION_OUTPUT_PATH_COLUMN_NAME,
    PROTONATION_STATUS_COLUMN_NAME,
    RANGE_COLUMN_NAME,
    RED_HEX,
    RETAIN_ALPHAFOLD_VARIANT_COLUMN_NAME,
    RETAIN_GAPS_VARIANT_COLUMN_NAME,
    RETAIN_MODELLER_VARIANT_COLUMN_NAME,
    REVIEW_FILL,
    ROW_STRIPE_FILL,
    SEMANTIC_COLOR_COLUMN_NAME_SET,
    SEQUENCE_ALIGNMENT_DONE_COLUMN_NAME,
    SKIP_FILL,
    STATUS_COLOR_COLUMN_NAME_SET,
    STATUS_GROUP_COLUMN_NAME_SET,
    STATUS_GROUP_FILL,
    STRONG_REVIEW_FILL,
    SUCCESS_FILL,
    TITLE_ALIGNMENT,
    TITLE_FILL,
    TITLE_FONT,
    TITLE_FONT_SIZE,
    UNIPROT_ID_COLUMN_NAME,
    VARIANT_POLICY_COLUMN_NAME,
    WHITE_BOLD_FONT,
    WHITE_HEX,
    WORKSHEET_LAYOUTS,
    STATE_COLUMN_NAME_LIST,
    STATUS_FAILED,
    STATUS_REQUIRED,
    STATUS_SKIPPED,
    STATUS_SUCCESS,
    STATUS_WARNING,
    STEP_STATUS_COLUMN_NAME_LIST,
)
from stack_protein_preparation._xlsx_formatters import (  # noqa: F401 (re-exported)
    _set_white_bold,
    _set_click_hyperlink,
    _set_row_height,
    apply_base_data_style,
    apply_group_body_shading,
    apply_metal_geometry_pair_color,
    apply_row_striping,
    apply_semantic_yes_no_color,
    apply_status_cell_color,
    autosize_worksheet_columns,
    set_worksheet_view_options,
    style_header_row,
)

# Column names, palette constants, and style objects are defined in
# _xlsx_columns.py and imported above. Formatting functions are in
# _xlsx_formatters.py.



# ---------------------------------------------------------------------------
# Record normalization and derived values
# ---------------------------------------------------------------------------


def get_excel_column_order(
    protein_record_list: list[dict[str, Any]] | None = None,
) -> list[str]:
    """Return the full export column order including known extra fields.

    The JSON state is defined centrally, but some recent FRUTON stages may add
    extra diagnostic keys before ``pipeline_state.py`` is extended. The XLSX
    exporter keeps those values visible in ``full_state`` instead of discarding
    them. Derived display columns are appended first so they can be reused by
    compact worksheets without changing the persisted JSON schema.
    """

    column_order = list(STATE_COLUMN_NAME_LIST)

    for column_name in OPTIONAL_KNOWN_EXTRA_COLUMN_NAME_LIST:
        if column_name not in column_order:
            column_order.append(column_name)

    if protein_record_list:
        extra_column_names = sorted(
            {
                str(column_name)
                for protein_record in protein_record_list
                for column_name in protein_record
                if str(column_name) not in column_order
            }
        )
        column_order.extend(extra_column_names)

    for column_name in DERIVED_COLUMN_NAME_LIST:
        if column_name not in column_order:
            column_order.append(column_name)

    return column_order


def ensure_all_records_have_all_columns(
    protein_record_list: list[dict[str, Any]],
) -> None:
    """Ensure that every record has at least the core pipeline columns."""

    empty_record = create_empty_protein_record()

    for protein_record in protein_record_list:
        for column_name in empty_record:
            if column_name not in protein_record:
                protein_record[column_name] = ""


def normalize_record_values(
    protein_record_list: list[dict[str, Any]],
) -> list[dict[str, str]]:
    """Return normalized copies of protein records for export.

    The function preserves known state columns, optional diagnostic columns, and
    any extra keys present in the incoming record dictionaries. It then enriches
    each record with XLSX-only derived fields that summarize whether a prepared
    final model exists and which path should be treated as the main output. This
    keeps the workbook useful even when the JSON state evolves faster than the
    spreadsheet module.
    """

    normalized_record_list: list[dict[str, str]] = []
    column_order = get_excel_column_order(protein_record_list)

    for protein_record in protein_record_list:
        normalized_record: dict[str, str] = {}

        for column_name in column_order:
            raw_value = protein_record.get(column_name, "")
            normalized_value = str(raw_value).strip()

            if column_name == PDB_ID_COLUMN_NAME:
                normalized_value = normalized_value.upper()

            normalized_record[column_name] = normalized_value

        if not normalized_record[PDB_ID_COLUMN_NAME]:
            continue

        _add_derived_final_model_fields(normalized_record)
        normalized_record_list.append(normalized_record)

    return normalized_record_list


def create_unique_sorted_record_list(
    protein_record_list: list[dict[str, str]],
) -> list[dict[str, str]]:
    """Return unique records sorted by PDB ID."""

    record_by_pdb_id: dict[str, dict[str, str]] = {}

    for protein_record in protein_record_list:
        pdb_id = str(protein_record.get(PDB_ID_COLUMN_NAME, "")).strip().upper()
        if not pdb_id:
            continue

        copied_record = dict(protein_record)
        copied_record[PDB_ID_COLUMN_NAME] = pdb_id
        record_by_pdb_id[pdb_id] = copied_record

    return [record_by_pdb_id[pdb_id] for pdb_id in sorted(record_by_pdb_id.keys())]


def _add_derived_final_model_fields(protein_record: dict[str, str]) -> None:
    """Add spreadsheet-only final-model summary fields to one record."""

    final_path = _choose_final_model_path(protein_record)
    final_type = _choose_final_model_type(protein_record, final_path)

    protein_record[FINAL_MODEL_CREATED_COLUMN_NAME] = "yes" if final_path else "no"
    protein_record[FINAL_MODEL_TYPE_COLUMN_NAME] = final_type
    protein_record[FINAL_MODEL_PATH_COLUMN_NAME] = final_path


def _choose_final_model_path(protein_record: dict[str, str]) -> str:
    """Return the preferred final prepared model path, or an empty string."""

    for column_name in (
        PREPARED_STRUCTURE_OUTPUT_PATH_COLUMN_NAME,
        PREPARED_MODELLER_OUTPUT_PATH_COLUMN_NAME,
        PREPARED_ALPHAFOLD_OUTPUT_PATH_COLUMN_NAME,
        PREPARED_GAPS_OUTPUT_PATH_COLUMN_NAME,
    ):
        value = str(protein_record.get(column_name, "")).strip()
        if value:
            return value

    return ""


def _choose_final_model_type(
    protein_record: dict[str, str],
    final_path: str,
) -> str:
    """Return a compact label for the final model type."""

    if not final_path:
        return ""

    prepared_variant = str(
        protein_record.get(PREPARED_STRUCTURE_VARIANT_COLUMN_NAME, "")
    ).strip()
    if prepared_variant:
        return prepared_variant

    available_models = str(protein_record.get(AVAILABLE_MODELS_COLUMN_NAME, "")).strip()
    if available_models:
        return available_models

    return "prepared"


def _record_has_final_model(protein_record: dict[str, str]) -> bool:
    return str(protein_record.get(FINAL_MODEL_CREATED_COLUMN_NAME, "")).lower() == "yes"


# ---------------------------------------------------------------------------
# Column filtering
# ---------------------------------------------------------------------------


def get_nonempty_column_order_for_subset(
    protein_record_list: list[dict[str, str]],
    base_column_order: list[str],
    always_keep_column_name_set: set[str] | None = None,
) -> list[str]:
    """Return only useful columns from one requested worksheet layout."""

    always_keep_column_name_set = always_keep_column_name_set or set()
    filtered_column_order: list[str] = []

    for column_name in base_column_order:
        if column_name in always_keep_column_name_set:
            filtered_column_order.append(column_name)
            continue

        has_nonempty_value = any(
            str(protein_record.get(column_name, "")).strip()
            for protein_record in protein_record_list
        )

        if has_nonempty_value:
            filtered_column_order.append(column_name)

    return filtered_column_order



# ---------------------------------------------------------------------------
# Logo helpers
# ---------------------------------------------------------------------------


def _resolve_logo_path(
    *,
    output_path: Path,
    logo_path: str | Path | None,
) -> Path | None:
    """Return the first available logo path for the overview header.

    The function keeps the public writer convenient for the FRUTON runner by
    accepting an optional explicit path while also checking sensible project
    locations. The most useful default is ``data/logo.png`` at the repository
    root, because the workbook is usually written to ``data/proteins``. Missing
    logos never fail the export; the sheet falls back to a clean text-only
    header so automated pipeline runs remain robust.
    """

    candidate_path_list: list[Path] = []

    if logo_path is not None:
        candidate_path_list.append(Path(logo_path).expanduser())

    output_path = Path(output_path).expanduser()

    # Repo-relative path: works from $LUSTRE when __file__ is on $STORE.
    # tests/src/stack_protein_preparation/ → repo root is 3 levels up.
    _here = Path(__file__).resolve().parent
    _repo_root = _here.parent.parent
    for filename in LOGO_FILENAME_CANDIDATE_LIST:
        candidate_path_list.extend(
            [
                _repo_root / "data" / filename,          # repo's data/ — always on $STORE
                _repo_root / "data" / "assets" / filename,
                output_path.parent / filename,           # next to pipeline.xlsx
                output_path.parent.parent / filename,
                output_path.parent / "assets" / filename,
                Path.cwd() / "data" / filename,
                Path.cwd() / filename,
            ]
        )

    for candidate_path in candidate_path_list:
        try:
            resolved_path = candidate_path.resolve()
        except Exception:
            resolved_path = candidate_path

        if resolved_path.exists() and resolved_path.is_file():
            return resolved_path

    return None


def _add_logo_to_worksheet(
    worksheet,
    *,
    logo_path: Path | None,
) -> bool:
    """Place the FRUTON logo using the manually tuned Calc geometry.

    LibreOffice reports object position and size in inches, while XLSX stores
    drawing anchors in EMUs. A one-cell anchor with explicit EMU offsets keeps
    the logo in the top-left dashboard band without depending on column widths
    or row heights. The size intentionally follows the inspected Calc values:
    X=0.04 in, Y=0.00 in, W=0.74 in, H=0.84 in.
    """

    if logo_path is None:
        return False

    try:
        # Read bytes into memory first so openpyxl embeds the data directly
        # into the xlsx archive rather than re-opening the path at save time.
        # This guarantees the logo is present when the file is opened on any
        # machine (e.g. downloaded from CESGA $LUSTRE to a local PC).
        logo_bytes = io.BytesIO(Path(logo_path).read_bytes())
        logo_image = OpenpyxlImage(logo_bytes)
    except Exception:
        return False

    marker = AnchorMarker(
        col=0,
        colOff=inch_to_EMU(LOGO_OFFSET_X_IN),
        row=0,
        rowOff=inch_to_EMU(LOGO_OFFSET_Y_IN),
    )
    size = XDRPositiveSize2D(
        cx=inch_to_EMU(LOGO_WIDTH_IN),
        cy=inch_to_EMU(LOGO_HEIGHT_IN),
    )
    logo_image.anchor = OneCellAnchor(_from=marker, ext=size)
    worksheet.add_image(logo_image)
    return True

# ---------------------------------------------------------------------------
# KPI helpers
# ---------------------------------------------------------------------------


def _count_records_where(
    protein_record_list: list[dict[str, str]],
    column_name: str,
    value: str,
) -> int:
    return sum(
        1
        for protein_record in protein_record_list
        if protein_record.get(column_name, "").strip().lower() == value
    )


def _build_overview_kpi_list(
    protein_record_list: list[dict[str, str]],
) -> list[tuple[str, int]]:
    """Return compact top-row metrics for the overview worksheet."""

    total = len(protein_record_list)
    final_models = _count_records_where(
        protein_record_list,
        FINAL_MODEL_CREATED_COLUMN_NAME,
        "yes",
    )

    return [
        ("records", total),
        ("final models", final_models),
        ("missing final", total - final_models),
        (
            "metals",
            _count_records_where(protein_record_list, HAS_METALS_COLUMN_NAME, "yes"),
        ),
        (
            "metal ready",
            _count_records_where(protein_record_list, METALS_MODEL_READY_COLUMN_NAME, "yes"),
        ),
        (
            "nonstandard",
            _count_records_where(
                protein_record_list,
                HAS_NONSTANDARD_RESIDUES_COLUMN_NAME,
                "yes",
            ),
        ),
        (
            "gap cases",
            _count_records_where(protein_record_list, HAS_GAPS_COLUMN_NAME, "yes"),
        ),
    ]


# ---------------------------------------------------------------------------
# Worksheet writing
# ---------------------------------------------------------------------------


def _write_overview_title_and_kpis(
    worksheet,
    protein_record_list: list[dict[str, str]],
    column_count: int,
    *,
    logo_path: Path | None,
) -> None:
    """Write the overview header, gold band, and KPI strip.

    The merge pattern follows the manually adjusted Calc layout: ``A1:A2`` is
    the logo cell, ``B1:last`` is the title, ``B2:last`` is the subtitle, and
    ``A3:last`` is one continuous gold accent row. The title is ``SUMMARY``
    because the workbook summarizes a framework run, not a generic pipeline.
    """

    max_column = max(column_count, 14)
    last_column_letter = get_column_letter(max_column)

    for row_index in range(1, 3):
        for column_index in range(1, max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.fill = TITLE_FILL
            cell.border = HEADER_BORDER

    worksheet.merge_cells("A1:A2")
    logo_added = _add_logo_to_worksheet(worksheet, logo_path=logo_path)

    logo_cell = worksheet.cell(row=1, column=1)
    if not logo_added:
        logo_cell.value = "FRUTON"
        logo_cell.font = TITLE_FONT
        logo_cell.alignment = CENTER_ALIGNMENT

    worksheet.merge_cells(f"B1:{last_column_letter}1")
    title_cell = worksheet.cell(row=1, column=2)
    title_cell.value = "SUMMARY"
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = TITLE_ALIGNMENT
    title_cell.border = HEADER_BORDER

    worksheet.merge_cells(f"B2:{last_column_letter}2")
    subtitle_cell = worksheet.cell(row=2, column=2)
    subtitle_cell.value = "model availability · gaps · chemistry · preparation status"
    subtitle_cell.fill = TITLE_FILL
    subtitle_cell.font = Font(name=FONT_NAME, size=10, color=WHITE_HEX)
    subtitle_cell.alignment = TITLE_ALIGNMENT
    subtitle_cell.border = HEADER_BORDER

    worksheet.merge_cells(f"A3:{last_column_letter}3")
    accent_cell = worksheet.cell(row=3, column=1)
    accent_cell.value = ""
    accent_cell.fill = STRONG_REVIEW_FILL
    accent_cell.border = HEADER_BORDER

    for column_index in range(1, max_column + 1):
        cell = worksheet.cell(row=4, column=column_index)
        cell.fill = KPI_VALUE_FILL
        cell.border = GRID_BORDER

    _set_row_height(worksheet, 1, 31.0)
    _set_row_height(worksheet, 2, 20.25)
    _set_row_height(worksheet, 3, 10.1)
    _set_row_height(worksheet, 4, 22)

    kpi_list = _build_overview_kpi_list(protein_record_list)
    start_column = 1

    for label, value in kpi_list:
        label_cell = worksheet.cell(row=4, column=start_column)
        value_cell = worksheet.cell(row=4, column=start_column + 1)

        label_cell.value = label
        value_cell.value = value

        label_cell.fill = KPI_LABEL_FILL
        label_cell.font = KPI_LABEL_FONT
        label_cell.alignment = CENTER_ALIGNMENT
        label_cell.border = GRID_BORDER

        value_cell.fill = KPI_VALUE_FILL
        value_cell.font = KPI_VALUE_FONT
        value_cell.alignment = CENTER_ALIGNMENT
        value_cell.border = GRID_BORDER

        if label == "missing final" and value:
            value_cell.font = PDB_MISSING_FONT

        start_column += 2


def _append_record_row(
    worksheet,
    protein_record: dict[str, str],
    column_order: list[str],
) -> None:
    row_value_list = [
        protein_record.get(column_name, "") for column_name in column_order
    ]
    worksheet.append(row_value_list)


def _style_record_rows(
    worksheet,
    protein_record_list: list[dict[str, str]],
    column_order: list[str],
    *,
    first_data_row_index: int,
) -> None:
    for offset, _protein_record in enumerate(protein_record_list):
        row_index = first_data_row_index + offset

        for column_index, column_name in enumerate(column_order, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell_value = "" if cell.value is None else str(cell.value).strip()

            if column_name == PDB_ID_COLUMN_NAME:
                cell.font = PDB_FONT
                cell.alignment = CENTER_ALIGNMENT

            if column_name in STATUS_COLOR_COLUMN_NAME_SET:
                apply_status_cell_color(cell, cell_value)

            if column_name in SEMANTIC_COLOR_COLUMN_NAME_SET:
                apply_semantic_yes_no_color(
                    cell=cell,
                    column_name=column_name,
                    value=cell_value,
                )
            if column_name in {N_GAPS_COLUMN_NAME, GAP_SIZES_COLUMN_NAME}:
                if cell_value and cell_value.lower() not in {"0", "none"}:
                    cell.fill = REVIEW_FILL


            if column_name in HYPERLINK_COLUMN_NAME_SET and cell_value:
                _set_click_hyperlink(cell, cell_value)


def _write_one_worksheet(
    workbook: Workbook,
    worksheet_name: str,
    protein_record_list: list[dict[str, str]],
    base_column_order: list[str],
    *,
    always_keep_column_name_set: set[str] | None = None,
    logo_path: Path | None = None,
) -> None:
    column_order = get_nonempty_column_order_for_subset(
        protein_record_list=protein_record_list,
        base_column_order=base_column_order,
        always_keep_column_name_set=always_keep_column_name_set,
    )

    worksheet = workbook.create_sheet(title=worksheet_name)

    if worksheet_name == "overview":
        header_row_index = 6
        first_data_row_index = 7
        _write_overview_title_and_kpis(
            worksheet=worksheet,
            protein_record_list=protein_record_list,
            column_count=len(column_order),
            logo_path=logo_path,
        )
        worksheet.append([])
        worksheet.append(column_order)
    else:
        header_row_index = 1
        first_data_row_index = 2
        worksheet.append(column_order)

    style_header_row(
        worksheet=worksheet,
        column_order=column_order,
        header_row_index=header_row_index,
    )

    for protein_record in protein_record_list:
        _append_record_row(
            worksheet=worksheet,
            protein_record=protein_record,
            column_order=column_order,
        )

    apply_base_data_style(
        worksheet=worksheet,
        first_data_row_index=first_data_row_index,
    )
    apply_group_body_shading(
        worksheet=worksheet,
        column_order=column_order,
        first_data_row_index=first_data_row_index,
    )
    _style_record_rows(
        worksheet=worksheet,
        protein_record_list=protein_record_list,
        column_order=column_order,
        first_data_row_index=first_data_row_index,
    )
    apply_metal_geometry_pair_color(
        worksheet=worksheet,
        protein_record_list=protein_record_list,
        column_order=column_order,
        first_data_row_index=first_data_row_index,
    )
    apply_row_striping(
        worksheet=worksheet,
        first_data_row_index=first_data_row_index,
    )
    mark_missing_model_pdb_ids(
        worksheet=worksheet,
        protein_record_list=protein_record_list,
        column_order=column_order,
        first_data_row_index=first_data_row_index,
    )

    for row_index in range(first_data_row_index, worksheet.max_row + 1):
        _set_row_height(worksheet, row_index, 18)

    autosize_worksheet_columns(worksheet)

    last_column_letter = get_column_letter(max(1, len(column_order)))
    auto_filter_range = (
        f"A{header_row_index}:{last_column_letter}{worksheet.max_row}"
    )
    set_worksheet_view_options(
        worksheet=worksheet,
        freeze_cell=f"A{first_data_row_index}",
        auto_filter_range=auto_filter_range,
    )


# ---------------------------------------------------------------------------
# Public writer
# ---------------------------------------------------------------------------


def write_pipeline_to_xlsx(
    protein_record_list: list[dict[str, Any]],
    output_path: Path,
    logo_path: str | Path | None = None,
) -> None:
    """Write pipeline records to a polished multi-sheet XLSX file."""

    working_record_list = [
        dict(protein_record) for protein_record in protein_record_list
    ]

    ensure_all_records_have_all_columns(working_record_list)
    normalized_record_list = normalize_record_values(working_record_list)
    unique_sorted_record_list = create_unique_sorted_record_list(normalized_record_list)

    # Convert absolute path values to paths relative to the xlsx output directory.
    # This makes the workbook portable when the results folder is downloaded.
    _xlsx_dir = Path(output_path).parent
    for _rec in unique_sorted_record_list:
        for _col, _val in list(_rec.items()):
            if not isinstance(_val, str):
                continue
            _stripped = _val.strip()
            if not _stripped or _stripped.startswith("."):
                continue
            # Heuristic: treat as path if it looks like an absolute filesystem path
            if _stripped.startswith("/") or (_stripped[1:3] in (":\\", ":/") if len(_stripped) > 2 else False):
                try:
                    _abs = Path(_stripped)
                    if _abs.is_absolute():
                        try:
                            _rel = _abs.relative_to(_xlsx_dir)
                            _rec[_col] = str(_rel)
                        except ValueError:
                            pass  # path not under xlsx_dir — keep as-is
                except Exception:
                    pass

    output_path = Path(output_path)
    resolved_logo_path = _resolve_logo_path(
        output_path=output_path,
        logo_path=logo_path,
    )

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for worksheet_name, base_column_order in WORKSHEET_LAYOUTS:
        if worksheet_name == "full_state":
            base_column_order = get_excel_column_order(unique_sorted_record_list)

        always_keep = {PDB_ID_COLUMN_NAME}

        if worksheet_name == "overview":
            always_keep |= {
                UNIPROT_ID_COLUMN_NAME,
                RANGE_COLUMN_NAME,
                FINAL_MODEL_CREATED_COLUMN_NAME,
                FINAL_MODEL_TYPE_COLUMN_NAME,
                AVAILABLE_MODELS_COLUMN_NAME,
            }

        if worksheet_name in {"gaps_and_variants", "prepared_outputs"}:
            always_keep |= {
                FINAL_MODEL_CREATED_COLUMN_NAME,
                AVAILABLE_MODELS_COLUMN_NAME,
            }

        _write_one_worksheet(
            workbook=workbook,
            worksheet_name=worksheet_name,
            protein_record_list=unique_sorted_record_list,
            base_column_order=base_column_order,
            always_keep_column_name_set=always_keep,
            logo_path=resolved_logo_path if worksheet_name == "overview" else None,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)